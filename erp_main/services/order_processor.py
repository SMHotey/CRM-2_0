import re
from collections import Counter
from openpyxl import load_workbook
from django.core.exceptions import ValidationError


class OrderProcessor:
    """Сервис для обработки заказов из файлов"""

    @staticmethod
    def validate_file_header(sheet):
        """Проверка заголовка файла"""
        return sheet.cell(row=1, column=3).value == "Бланк №"

    @staticmethod
    def process_file_data(sheet):
        """Обработка данных из файла Excel"""
        cur_row, cur_column = 9, 15
        while sheet.cell(row=cur_row, column=cur_column).value != 'шт.':
            cur_row += 1
        max_row = cur_row

        seq = [1, 2, 3, 4, 5, 6, 9, 10, 11, 12, 13, 14, 15, 7, 8]
        positions = []
        line = []

        for row in range(8, max_row):
            if sheet.cell(row=row, column=2).value:
                if line:
                    positions.append(line)
                line = [sheet.cell(row=row, column=column).value for column in seq]
            else:
                line.extend([sheet.cell(row=row, column=7).value, sheet.cell(row=row, column=8).value])

        if line:
            positions.append(line)

        return positions

    @staticmethod
    def get_product_kind(name):
        """Определение типа продукта"""
        kind_mapping = {
            'дверь': 'door', 'люк': 'hatch', 'ворота': 'gate',
            'калитка': 'door', 'фрамуга': 'transom'
        }
        return next((value for key, value in kind_mapping.items()
                     if re.search(key, name, re.IGNORECASE)), None)

    @staticmethod
    def get_product_type(name):
        """Определение вида продукта"""
        type_mapping = {
            'ei-60': 'ei-60', 'eis-60': 'eis-60', 'eiws-60': 'eiws-60',
            'тех': 'tech', 'ревиз': 'revision'
        }
        return next((value for key, value in type_mapping.items()
                     if re.search(key, name, re.IGNORECASE)), None)

    @staticmethod
    def count_glass_data(glass_data):
        """Подсчет данных о стеклах"""
        counted_glass = dict(Counter(list(zip(glass_data[::2], glass_data[1::2]))))
        counted_glass.pop((None, None), None)
        return counted_glass