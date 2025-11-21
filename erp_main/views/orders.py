from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.views.generic import FormView
from django.core.exceptions import ValidationError, PermissionDenied
from django.db.models import Q
from openpyxl import load_workbook
import json
import logging
import re
from django.contrib import messages

from ..models import Order, OrderItem, Organization, InternalLegalEntity, GlassInfo, OrderChangeHistory
from ..forms import OrderForm, OrderFileForm
from .mixins import UserAccessMixin
from .permissions import (  # Импорт из нового файла permissions.py
    get_user_role_from_request,
    has_permission_for_action,
    can_view_order,
    can_edit_order_detail,
    can_modify_order_item, ajax_permission_required
)
from ..services.order_processor import OrderProcessor

logger = logging.getLogger(__name__)


class OrderUploadView(UserAccessMixin, FormView):
    template_name = 'order_upload.html'
    form_class = OrderForm
    processor = OrderProcessor()

    # Определяем требуемые роли для доступа к загрузке заказов
    required_roles = ['admin', 'director', 'manager', 'production']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_role = self.get_user_role()

        # Сохраняем предыдущую логику доступа к организациям
        if self.request.user.is_superuser or user_role in ['admin', 'director']:
            context['organizations'] = Organization.objects.all()
        else:
            context['organizations'] = Organization.objects.filter(user=self.request.user)

        context['internal_legal_entities'] = InternalLegalEntity.objects.all()
        context['user_role'] = user_role  # Передаем роль в шаблон
        return context

    def get_form_kwargs(self):
        """Передаем текущего пользователя в форму"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Проверяем права доступа перед обработкой формы
        user_role = self.get_user_role()
        if user_role not in self.required_roles:
            raise PermissionDenied("У вас недостаточно прав для загрузки заказов")

        uploaded_file = form.cleaned_data.get('order_file')
        order_id = self.kwargs.get('order_id')

        # Для нового заказа файл обязателен
        if not order_id and not uploaded_file:
            form.add_error('order_file', 'Пожалуйста, загрузите файл.')
            return self.form_invalid(form)

        if order_id:
            # Извлечение существующего заказа из базы
            order = get_object_or_404(Order, pk=order_id)

            # Проверяем права на редактирование заказа
            if not self._can_edit_order(order):
                raise PermissionDenied("У вас недостаточно прав для редактирования этого заказа")

            old_file = order.order_file  # старый файл заказа для сохранения в истории изменений
            is_update = True
        else:
            # Создание нового заказа на основе данных из формы
            order = Order()
            is_update = False
            order.invoice = form.cleaned_data.get('invoice')
            order.due_date = form.cleaned_data.get('due_date')
            order.comment = form.cleaned_data.get('comment', '')

        # Обновление поля order_file только если файл был загружен
        if uploaded_file:
            order.order_file = uploaded_file

            # Проверка файла
            try:
                wb = load_workbook(uploaded_file)
                if not self.processor.validate_file_header(wb.active):
                    form.add_error('order_file', 'Выберите правильный файл заказа')
                    return self.form_invalid(form)
            except Exception as e:
                form.add_error('order_file', 'Ошибка загрузки файла: ' + str(e))
                return self.form_invalid(form)

        # Сохраняем объект Order
        order.save()

        # Обработка загрузки файла и обновление позиций заказа
        if uploaded_file:
            wb = load_workbook(uploaded_file)
            new_positions = self.processor.process_file_data(wb.active)

            if is_update:
                self._update_order_items(order, new_positions, old_file)
            else:
                self._create_order_items(order, new_positions)

        return redirect(reverse('order_detail', args=[order.id]))

    def _can_edit_order(self, order):
        """Проверяет, может ли пользователь редактировать заказ"""
        user_role = self.get_user_role()

        # Админы и директоры могут редактировать все заказы
        if user_role in ['admin', 'director']:
            return True

        # Менеджеры и производство могут редактировать только свои заказы
        if user_role in ['manager', 'production']:
            return order.invoice.organization.user == self.request.user

        return False

    def _update_order_items(self, order, new_positions, old_file):
        """Обновление позиций существующего заказа"""
        current_items = {item.position_num: item for item in
                         OrderItem.objects.filter(order=order).prefetch_related('glasses')}
        changes_made = []

        for data in new_positions:
            n_num = str(data[0])
            name = data[1]
            n_kind = self.processor.get_product_kind(name)
            n_type = self.processor.get_product_type(name)
            n_construction = 'NK' if re.search('-м', name.lower()) else 'SK'

            new_item_data = {
                'p_kind': n_kind,
                'p_type': n_type,
                'p_construction': n_construction,
                'p_height': data[2],
                'p_width': data[3],
                'p_active_trim': data[4],
                'p_open': data[5],
                'p_platband': data[6],
                'p_furniture': data[7],
                'p_door_closer': data[8],
                'p_step': data[9],
                'p_ral': str(data[10]),
                'p_quantity': data[11],
                'p_comment': data[12],
                'p_glass': self.processor.count_glass_data(data[13:]),
            }

            if n_num in current_items:
                first = 0
                current_item = current_items[n_num]
                for field, new_value in new_item_data.items():
                    old_value = getattr(current_item, field, None)

                    # Handle glass changes separately
                    if field == 'p_glass':
                        # Get existing glass info from database
                        old_glass = {(g.height, g.width): g.quantity for g in current_item.glasses.all()}
                        new_glass = new_value

                        # Check if glass has changed
                        if old_glass != new_glass:
                            if first == 0:
                                changes_made.append(f'<br> поз. {n_num}: изменено стекло ')
                                first = 1

                            # Format old glass info
                            old_glass_str = ", ".join(
                                [f"{h}x{w} ({q} шт.)" for (h, w), q in old_glass.items()]) if old_glass else "нет"
                            # Format new glass info
                            new_glass_str = ", ".join(
                                [f"{h}x{w} ({q} шт.)" for (h, w), q in new_glass.items()]) if new_glass else "нет"

                            changes_made.append(f"с \"{old_glass_str}\" на \"{new_glass_str}\";")

                            # Update p_glass field
                            setattr(current_item, 'p_glass', str(new_glass))
                    else:
                        if str(old_value) != str(new_value):
                            field_name = OrderItem._meta.get_field(field).verbose_name
                            if old_value:
                                if first == 0:
                                    changes_made.append(
                                        f'<br> поз. {n_num}:  {field_name} с "{old_value}" на "{new_value}"; ')
                                    first = 1
                                else:
                                    changes_made.append(f'{field_name} с "{old_value}" на "{new_value}";')
                            else:
                                changes_made.append(f'поз. {n_num}: добавлен {field_name} "{new_value}";')
                            setattr(current_item, field, new_value)

                if changes_made:
                    # Delete existing glass info before saving new one
                    current_item.glasses.all().delete()
                    current_item.save()

                    # Save new glass info
                    for (height, width), quantity in new_item_data['p_glass'].items():
                        if height and width:  # Only create records for valid dimensions
                            GlassInfo.objects.create(
                                order_items=current_item,
                                height=height,
                                width=width,
                                quantity=quantity
                            )
            else:
                new_item = OrderItem(order=order, position_num=n_num, **new_item_data)
                new_item.p_status = 'in_query'
                new_item.save()

                # Save glass info for new item
                for (height, width), quantity in new_item_data['p_glass'].items():
                    if height and width:
                        GlassInfo.objects.create(
                            order_items=new_item,
                            height=height,
                            width=width,
                            quantity=quantity
                        )

        if changes_made:
            comment = str(changes_made).replace('[', '').replace(']', '').replace("'", "").replace(",", "").strip()[5::]
            order.save()
            add_changes = OrderChangeHistory(order=order, order_file=old_file, changed_by=self.request.user,
                                             comment=comment)
            add_changes.save()

            if order.order_file:
                try:
                    wb = load_workbook(order.order_file.path)
                    sheet = wb.active
                    sheet['K3'] = comment.replace('<br>', '')
                    wb.save(order.order_file.path)
                except Exception as e:
                    logger.error(f"Ошибка при записи комментария в файл: {e}")

    def _create_order_items(self, order, new_positions):
        """Создание новых позиций заказа"""
        for data in new_positions:
            n_num = data[0]
            name = data[1]
            n_kind = self.processor.get_product_kind(name)
            n_type = self.processor.get_product_type(name)
            n_construction = 'NK' if re.search('-м', name.lower()) else 'SK'

            new_item_data = {
                'p_kind': n_kind,
                'p_type': n_type,
                'p_construction': n_construction,
                'p_height': data[2],
                'p_width': data[3],
                'p_active_trim': data[4],
                'p_open': data[5],
                'p_platband': data[6],
                'p_furniture': data[7],
                'p_door_closer': data[8],
                'p_step': data[9],
                'p_ral': data[10],
                'p_quantity': data[11],
                'p_comment': data[12],
                'p_glass': self.processor.count_glass_data(data[13:]),
            }

            new_item = OrderItem(order=order, position_num=n_num, **new_item_data)
            new_item.p_status = 'in_query'
            new_item.save()

            # Сохраняем информацию о стеклах
            self._save_glass_info(new_item, new_item_data['p_glass'])

    def _save_glass_info(self, new_item, counted_glass):
        """Сохраняет информацию о стеклах для нового элемента заказа"""
        for (height, width), quantity in counted_glass.items():
            if height and width:
                GlassInfo.objects.create(
                    order_items=new_item,
                    height=height,
                    width=width,
                    quantity=quantity
                )

    def form_invalid(self, form):
        """Обработка невалидной формы"""
        organizations = Organization.objects.all()
        return self.render_to_response(self.get_context_data(form=form, organizations=organizations))


@login_required
def orders_list(request):
    """Список заказов с учетом ролей пользователей"""
    orders = []
    source = request.GET.get('source')
    user_role = get_user_role_from_request(request)

    # Сохраняем предыдущую логику фильтрации заказов
    if request.user.is_staff or user_role in ['admin', 'director', 'production', 'logistic']:
        orders = Order.objects.all().order_by('-id')
    elif Order.objects.filter(invoice__organization__user=request.user):
        orders = Order.objects.all().order_by('-id')
    else:
        # Для пользователей без специальных прав показываем только их заказы
        orders = Order.objects.filter(invoice__organization__user=request.user).order_by('-id')

    if source:
        orders = Order.objects.filter(invoice__organization=source).order_by('-id')

    paginator = Paginator(orders, 20)
    page_number = request.GET.get('page')
    orders_page = paginator.get_page(page_number)

    # Передаем роль пользователя в контекст
    context = {
        'orders': orders_page,
        'user_role': user_role
    }

    return render(request, 'orders_list.html', context)


@login_required
def order_detail(request, order_id):
    """Детали заказа с проверкой прав доступа"""
    order = get_object_or_404(Order, id=order_id)
    user_role = get_user_role_from_request(request)

    # Проверяем права доступа к заказу
    if not can_view_order(request.user, user_role, order):
        messages.error(request, "У вас недостаточно прав для просмотра этого заказа")
        return redirect('orders_list')

    changes = order.changes.all()

    # Отфильтрованные OrderItem, где статус не равен 'changed'
    filtered_items = order.items.exclude(p_status__in=['changed'])

    if request.method == 'POST':
        # Проверяем права на редактирование заказа
        if not can_edit_order_detail(request.user, user_role, order):
            messages.error(request, "У вас недостаточно прав для редактирования этого заказа")
            return redirect('order_detail', order_id=order.id)

        form = OrderFileForm(request.POST, request.FILES)
        if form.is_valid():
            order.order_file = form.cleaned_data['order_file']
            order.save()
            messages.success(request, "Файл заказа успешно обновлен")
            return redirect('order_detail', order_id=order.id)

    context = {
        'order': order,
        'filtered_items': filtered_items,
        'changes': changes,
        'user_role': user_role,  # Передаем роль в шаблон
    }

    return render(request, 'order_detail.html', context)


@require_POST
@login_required
def update_order_item_status(request):
    """Обновление статусов позиций заказа с проверкой прав"""
    try:
        data = json.loads(request.body)
        updates = data.get('updates', {})
        user_role = get_user_role_from_request(request)

        for item_id in updates.keys():
            order_item = get_object_or_404(OrderItem, id=item_id)

            # Проверяем права на изменение позиции заказа
            if not can_modify_order_item(request.user, user_role, order_item):
                return JsonResponse({
                    'status': 'error',
                    'message': 'У вас недостаточно прав для изменения этой позиции заказа'
                }, status=403)

            new_data = updates[item_id]

            # Сохраняем предыдущую логику обновления статусов
            order_item.p_status = new_data['status']
            order_item.workshop = new_data['workshop']
            if order_item.workshop == '2' and new_data['path'] != 'order_detail':
                order_item.p_status = 'stopped'
            if ((order_item.p_status == 'stopped' or order_item.p_status == 'canceled') and
                    new_data['path'] != 'order_detail'):
                order_item.workshop = '2'
            order_item.save()

        return JsonResponse({'status': 'success', 'message': 'Информация обновлена'})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Неверный формат'}, status=400)
    except Exception as e:
        logger.exception("Error updating order items' statuses")
        return JsonResponse({'status': 'error', 'message': 'An error occurred while processing your request.'},
                            status=500)


@require_POST
@login_required
def update_workshop(request, order_id):
    """Обновление статуса заказа с проверкой прав доступа"""
    try:
        data = json.loads(request.body)
        action = data.get('action')
        user_role = get_user_role_from_request(request)

        # Проверяем существование заказа
        order_items = OrderItem.objects.filter(order_id=order_id)
        if not order_items.exists():
            return JsonResponse({'success': False, 'error': 'Order not found'}, status=404)

        # Проверяем права на изменение заказа
        first_item = order_items.first()
        if not can_modify_order_item(request.user, user_role, first_item):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)

        # Получаем первый элемент заказа для определения текущего статуса
        current_status = first_item.p_status
        current_workshop = first_item.workshop

        # Проверяем права доступа для конкретного действия
        if not has_permission_for_action(user_role, current_status, current_workshop, action):
            return JsonResponse({'success': False, 'error': 'Permission denied for this action'}, status=403)

        # Обрабатываем действие
        result = process_order_action(order_items, action, request.user)

        if result['success']:
            # Записываем в историю изменений
            add_changes = OrderChangeHistory(
                order_id=order_id,
                changed_by=request.user,
                comment=result['comment']
            )
            add_changes.save()

            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': result['error']}, status=400)

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# Вспомогательные функции для проверки прав
# def _can_view_order(user, user_role, order):
#     """Проверяет, может ли пользователь просматривать заказ"""
#     if user_role in ['admin', 'director']:
#         return True
#     return order.invoice.organization.user == user
#
#
# def _can_edit_order_detail(user, user_role, order):
#     """Проверяет, может ли пользователь редактировать детали заказа"""
#     if user_role in ['admin', 'director']:
#         return True
#     if user_role in ['manager', 'production']:
#         return order.invoice.organization.user == user
#     return False
#
#
# def _can_modify_order_item(user, user_role, order_item):
#     """Проверяет, может ли пользователь изменять позицию заказа"""
#     if user_role in ['admin', 'director']:
#         return True
#     if user_role in ['manager', 'production', 'logistic']:
#         return order_item.order.invoice.organization.user == user
#     return False


def process_order_action(order_items, action, user):
    """Обрабатываем действие над заказом (сохраняем предыдущую логику)"""

    status_mapping = {
        'start_1': 'product',
        'start_3': 'product',
        'stop': 'stopped',
        'ready': 'ready',
        'cancel': 'canceled',
        'to_queue': 'in_query',
        'ship': 'shipped',
        'switch_1': 'product',
        'switch_3': 'product'
    }

    workshop_mapping = {
        'start_1': '1',
        'switch_1': '1',
        'start_3': '3',
        'switch_3': '3'
    }

    comment_mapping = {
        'start_1': 'заказ запущен в цехе №1',
        'start_3': 'заказ запущен в цехе №3',
        'stop': 'заказ остановлен',
        'ready': 'заказ готов',
        'cancel': 'заказ отменен',
        'to_queue': 'заказ возвращен в очередь',
        'ship': 'заказ отгружен',
        'switch_1': 'заказ переведен в цех №1',
        'switch_3': 'заказ переведен в цех №3'
    }

    new_status = status_mapping.get(action)
    new_workshop = workshop_mapping.get(action)
    comment = comment_mapping.get(action, 'статус заказа изменен')

    if not new_status:
        return {'success': False, 'error': 'Invalid action'}

    # Обновляем статус
    update_data = {'p_status': new_status}
    if new_workshop:
        update_data['workshop'] = new_workshop

    order_items.update(**update_data)

    return {'success': True, 'comment': comment}